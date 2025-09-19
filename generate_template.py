import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def create_template(output_path: str = "template.xlsx", 
                   accounts: Optional[Dict[str, List[Dict]]] = None) -> str:
    """
    Generate an Excel template for accounting with multiple accounts and dynamic columns.
    
    The template includes sample transactions with Indonesian column headers and supports multiple accounts.
    Each Penerimaan (income) and Pengeluaran (expense) category will be in its own column.
    
    Args:
        output_path: Path where the template will be saved
        accounts: Optional dictionary with account names as keys and lists of transactions as values.
                 If not provided, default sample data will be used.
                 
    Returns:
        str: Path to the generated template file
    """
    # Use provided accounts or create sample data if none provided
    if accounts is None:
        today = datetime.now()
        
        # Helper function to create a transaction entry
        def create_tx(days_ago: int, description: str, 
                     penerimaan: Optional[Dict] = None, 
                     pengeluaran: Optional[Dict] = None) -> Dict:
            """Helper to create a transaction dictionary"""
            return {
                "Tanggal": (today - timedelta(days=days_ago)).strftime("%Y-%m-%d"),
                "Uraian": description,
                "Penerimaan": penerimaan or {},
                "Pengeluaran": pengeluaran or {},
                "Jumlah": 0  # Will be calculated later
            }
        
        # Sample account data with more realistic transactions
        accounts_data = {
            "Kas": [
                create_tx(7, "Saldo Awal", {"Modal": 5000000}, {}),
                create_tx(6, "Pembayaran Piutang", {"Piutang": 2500000}, {}),
                create_tx(5, "Pembelian Bahan Baku", {}, {"Bahan Baku": 1750000}),
                create_tx(4, "Pendapatan Jasa", {"Jasa": 3200000}, {}),
                create_tx(3, "Biaya Listrik", {}, {"Listrik": 450000}),
                create_tx(2, "Pendapatan Lainnya", {"Lainnya": 1250000}, {}),
                create_tx(1, "Gaji Karyawan", {}, {"Gaji": 3000000})
            ],
            "Bank": [
                create_tx(7, "Setoran Awal", {"Setoran": 10000000}, {}),
                create_tx(5, "Pembayaran Supplier", {}, {"Supplier": 4500000}),
                create_tx(3, "Penerimaan Transfer", {"Transfer": 6500000}, {})
            ],
            "E-Wallet": [
                create_tx(7, "Saldo Awal", {"Saldo Awal": 1000000}, {}),
                create_tx(5, "Top Up", {"Top Up": 500000}, {}),
                create_tx(3, "Pembayaran Online", {}, {"Belanja Online": 750000}),
                create_tx(1, "Isi Ulang Pulsa", {}, {"Pulsa": 100000})
            ]
        }
        
        # Calculate running balances for each account
        for account_name, transactions in accounts_data.items():
            running_balance = 0
            for tx in transactions:
                income = sum(tx["Penerimaan"].values())
                expense = sum(tx["Pengeluaran"].values())
                running_balance += income - expense
                tx["Jumlah"] = running_balance
    else:
        accounts_data = accounts
    
    # If output_path is just a filename, save in current working directory
    output_path = Path(output_path)
    if not output_path.parent.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # If output_path is a directory, create a filename with timestamp
    if output_path.is_dir() or output_path.suffix.lower() != '.xlsx':
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"accounting_template_{timestamp}.xlsx"
        output_path = output_path / output_filename
    
    # Create a new Excel writer with openpyxl engine
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Create a worksheet for each account
        for account_name, transactions in accounts_data.items():
            if not transactions:
                continue
                
            # Find all unique column names for Penerimaan and Pengeluaran across all transactions
            all_penerimaan = set()
            all_pengeluaran = set()
            
            for tx in transactions:
                all_penerimaan.update(tx["Penerimaan"].keys())
                all_pengeluaran.update(tx["Pengeluaran"].keys())
            
            # Create a list to hold all rows
            rows = []
            
            for tx in transactions:
                row = {
                    "Tanggal": tx["Tanggal"],
                    "Uraian": tx["Uraian"],
                }
                
                # Add Penerimaan columns (income)
                for col in sorted(all_penerimaan):
                    row[f"Penerimaan_{col}"] = tx["Penerimaan"].get(col, 0)
                
                # Add Pengeluaran columns (expenses)
                for col in sorted(all_pengeluaran):
                    row[f"Pengeluaran_{col}"] = tx["Pengeluaran"].get(col, 0)
                
                # Add Jumlah (balance) and Saldo Berjalan (running balance)
                row["Jumlah"] = tx["Jumlah"]
                row["Saldo Berjalan"] = tx["Jumlah"]  # Initially same as Jumlah
                
                rows.append(row)
            
            # Create DataFrame and handle empty data case
            if not rows:
                # Create empty DataFrame with expected columns
                columns = ["Tanggal", "Uraian"]
                columns.extend([f"Penerimaan_{col}" for col in sorted(all_penerimaan)])
                columns.extend([f"Pengeluaran_{col}" for col in sorted(all_pengeluaran)])
                columns.extend(["Jumlah", "Saldo Berjalan"])
                df = pd.DataFrame(columns=columns)
            else:
                df = pd.DataFrame(rows)
            
            # Ensure consistent column order
            columns_order = ["Tanggal", "Uraian"]
            columns_order.extend(sorted([col for col in df.columns if col.startswith("Penerimaan_")]))
            columns_order.extend(sorted([col for col in df.columns if col.startswith("Pengeluaran_")]))
            columns_order.extend(["Jumlah", "Saldo Berjalan"])
            
            # Reorder and filter columns
            df = df.reindex(columns=[col for col in columns_order if col in df.columns])
            
            # Write to Excel with sheet name = account name (truncate if too long)
            sheet_name = account_name[:31]  # Excel sheet name limit
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Auto-adjust column widths and apply number formatting
            worksheet = writer.sheets[sheet_name]
            
            # Define number format for currency (using Indonesian format)
            currency_format = '#,##0_);(#,##0)'  # Positive numbers in black, negative in red
            
            for idx, column in enumerate(df.columns):
                # Set column width (wider for description, narrower for dates/numbers)
                if column == 'Uraian':
                    width = 40  # Wider for descriptions
                elif column == 'Tanggal':
                    width = 12
                elif column == 'Jumlah':
                    width = 15
                else:
                    width = 18  # Wider for monetary values
                
                # Apply number format to numeric columns (all except Tanggal and Uraian)
                if column not in ['Tanggal', 'Uraian']:
                    for cell in worksheet[get_column_letter(idx + 1) + '2':get_column_letter(idx + 1) + str(len(df) + 1)]:
                        cell[0].number_format = currency_format
                
                # Set column width
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = width
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
            
            # Add filters to the header row
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Format header row with different colors for different column types
            default_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            penerimaan_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")  # Light green
            pengeluaran_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
            
            # Apply header formatting based on column type
            for idx, cell in enumerate(worksheet[1], 1):
                column_title = cell.value
                if str(column_title).startswith('Penerimaan_'):
                    cell.fill = penerimaan_fill
                elif str(column_title).startswith('Pengeluaran_'):
                    cell.fill = pengeluaran_fill
                else:
                    cell.fill = default_fill
                
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    print(f"Template created with accounts: {', '.join(accounts_data.keys())}")
    return str(output_path.absolute())

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate an accounting template Excel file.')
    parser.add_argument('--output', '-o', type=str, default='template.xlsx',
                       help='Output file path (default: template.xlsx)')
    parser.add_argument('--list-accounts', action='store_true',
                       help='List available accounts in the template')
    
    args = parser.parse_args()
    
    if args.list_accounts:
        print("Available accounts in the template:")
        print("- Kas (Cash)")
        print("- Bank (Bank Account)")
        print("- E-Wallet (Digital Wallet)")
    else:
        path = create_template(args.output)
        print(f"Template created at: {path}")
