from fastapi import FastAPI, UploadFile, File, HTTPException, Response, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import pandas as pd
import io
from pathlib import Path
from datetime import datetime
import tempfile
import os
import generate_template
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# Initialize FastAPI app
app = FastAPI(title="Accounting Helper API")

# Allowed origins - update with your production domain
ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:3001",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:3001",
    "http://localhost:8000",
    "https://*.onrender.com",  # Allow Render deployment
    "https://*.vercel.app",    # Allow Vercel if you use it for frontend
    "https://*-*.vercel.app",  # Allow Vercel preview deployments
    "https://*-*-*.vercel.app" # Allow Vercel branch previews
]

# Enable CORS with dynamic origin handling
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=86400,  # 24 hours
)


# Add OPTIONS route for CORS preflight
@app.options("/api/save")
async def options_save(request: Request):
    return JSONResponse(
        content={"message": "OK"},
        headers={
            "Access-Control-Allow-Origin": request.headers.get("origin", "*"),
            "Access-Control-Allow-Methods": "POST, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type, Authorization",
            "Access-Control-Allow-Credentials": "true",
            "Access-Control-Max-Age": "86400",
        },
    )


# Mount static files for frontend if dist directory exists
if Path("dist").exists():
    app.mount("/static", StaticFiles(directory="dist"), name="static")
    print("Serving static files from 'dist' directory")
else:
    print("Note: 'dist' directory not found. Running in API-only mode.")


# Utility functions
def to_camel_case(snake_str: str) -> str:
    """Convert snake_case to camelCase."""
    components = snake_str.split("_")
    return components[0] + "".join(x.title() for x in components[1:])


# Pydantic models
class TransactionItem(BaseModel):
    """Represents a single transaction entry."""

    tanggal: str
    uraian: str
    penerimaan: Dict[str, float] = {}
    pengeluaran: Dict[str, float] = {}
    jumlah: float

    class Config:
        alias_generator = to_camel_case
        allow_population_by_field_name = True


class AccountData(BaseModel):
    """Represents an account with its transactions."""

    name: str
    id: Optional[str] = None
    transactions: List[TransactionItem] = []

    class Config:
        alias_generator = to_camel_case
        allow_population_by_field_name = True


class TemplateData(BaseModel):
    """Wrapper for account data."""

    accounts: List[AccountData]


# API Endpoints
@app.get("/")
async def root():
    """Root endpoint to check if the API is running."""
    return {"message": "Accounting Helper API is running", "status": "ok"}


@app.get("/api/health")
async def health_check(delay: int = 0):
    """
    Health check endpoint to monitor the backend status.
    This endpoint can be called by the frontend to keep the Koyeb instance awake.
    
    Args:
        delay: Optional delay in seconds before responding (for testing)
    """
    if delay > 0:
        import asyncio
        await asyncio.sleep(delay)
        
    return {
        "status": "ok",
        "timestamp": datetime.utcnow().isoformat(),
        "service": "accounting-helper-api"
    }


@app.get("/api/template")
async def download_template() -> Response:
    """Download the Excel template file with the new format."""
    temp_path = None
    try:
        print("\n=== Template Download Request ===")
        print("Generating template file...")

        # Create a temporary file with delete=False to manage it ourselves
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name

        # Generate the template file
        generate_template.create_template(output_path=temp_path)

        if not os.path.exists(temp_path):
            raise RuntimeError("Failed to generate template file")

        # Read the file content
        with open(temp_path, "rb") as f:
            file_content = f.read()

        # Create the response with the file content
        filename = f"accounting_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

        return response

    except Exception as e:
        print(f"Error in download_template: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Error generating template: {str(e)}"
        )
    finally:
        # Clean up the temporary file
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as e:
                print(f"Error cleaning up temp file: {e}")

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)) -> JSONResponse:
    """Process uploaded Excel file and return structured data."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Only Excel files (.xlsx, .xls) are supported"
        )

    try:
        contents = await file.read()
        xls = pd.ExcelFile(io.BytesIO(contents))
        accounts = _process_excel_sheets(xls)

        if not accounts:
            raise HTTPException(status_code=400, detail="No valid account sheets found")

        return JSONResponse(content={"accounts": accounts})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/api/save")
async def save_file(data: TemplateData, request: Request):
    temp_path = None
    try:
        print("\n=== Incoming Save Request ===")
        print(f"Request origin: {request.headers.get('origin')}")
        print(f"Request data type: {type(data)}")

        # Debug: Print account details
        if data.accounts:
            for i, account in enumerate(data.accounts):
                print(f"\nAccount {i + 1}:")
                print(f"  Name: {account.name}")
                print(f"  ID: {getattr(account, 'id', 'N/A')}")
                print(
                    f"  Number of transactions: {len(account.transactions) if hasattr(account, 'transactions') and account.transactions else 'No transactions'}"
                )
                if hasattr(account, "transactions") and account.transactions:
                    print(
                        "  First transaction:",
                        account.transactions[0].dict()
                        if account.transactions[0]
                        else "Empty transaction",
                    )

        # Create a temporary file with delete=False to manage it ourselves
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name

        # Generate the Excel file
        _create_excel_file(data, temp_path)

        # Read the file content
        with open(temp_path, "rb") as f:
            file_content = f.read()

        # Get the origin from the request headers
        origin = request.headers.get("origin", "*")

        # Create a response with the file content
        headers = {
            "Content-Disposition": f'attachment; filename="accounting_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"',
            "Access-Control-Allow-Origin": origin if any(origin.startswith(allowed.replace('*', '')) for allowed in ALLOWED_ORIGINS) else "*",
            "Access-Control-Allow-Credentials": "true",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }

        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except Exception as e:
        print(f"\n!!! Error in save_file: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        import traceback

        traceback.print_exc()

        # Get the origin from the request headers
        origin = request.headers.get("origin", "*")

        # Include CORS headers in error response
        headers = {
            "Access-Control-Allow-Origin": origin if any(origin.startswith(allowed.replace('*', '')) for allowed in ALLOWED_ORIGINS) else "*",
            "Access-Control-Allow-Credentials": "true",
        }

        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel file: {str(e)}",
            headers=headers,
        )
    finally:
        # Clean up the temporary file if it exists
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as e:
                print(f"Error cleaning up temp file: {e}")


# Helper functions
def _process_excel_sheets(xls: pd.ExcelFile) -> List[Dict[str, Any]]:
    """Process all sheets in the Excel file and return account data."""
    accounts = []
    required_columns = ["Tanggal", "Uraian", "Jumlah"]

    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name).fillna("")

            # Check for required columns
            if not all(col in df.columns for col in required_columns):
                continue

            # Get all Penerimaan and Pengeluaran columns
            penerimaan_cols = [
                col for col in df.columns if col.startswith("Penerimaan_")
            ]
            pengeluaran_cols = [
                col for col in df.columns if col.startswith("Pengeluaran_")
            ]

            transactions = _process_transactions(df, penerimaan_cols, pengeluaran_cols)
            accounts.append({"name": sheet_name, "transactions": transactions})
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {str(e)}")
            continue

    return accounts


def _process_transactions(
    df: pd.DataFrame, penerimaan_cols: List[str], pengeluaran_cols: List[str]
) -> List[Dict[str, Any]]:
    """Convert DataFrame rows to transaction dictionaries with dynamic columns."""
    transactions = []
    running_balance = 0.0

    for _, row in df.iterrows():
        try:
            # Calculate total penerimaan and pengeluaran for this transaction
            total_penerimaan = 0.0
            total_pengeluaran = 0.0
            
            # Process Penerimaan
            penerimaan = {}
            for col in penerimaan_cols:
                value = row[col]
                if pd.notna(value) and value != "" and float(value) != 0:
                    category = col.replace("Penerimaan_", "")
                    amount = float(value)
                    penerimaan[category] = amount
                    total_penerimaan += amount

            # Process Pengeluaran
            pengeluaran = {}
            for col in pengeluaran_cols:
                value = row[col]
                if pd.notna(value) and value != "" and float(value) != 0:
                    category = col.replace("Pengeluaran_", "")
                    amount = float(value)
                    pengeluaran[category] = amount
                    total_pengeluaran += amount

            # Calculate running balance
            running_balance = running_balance + total_penerimaan - total_pengeluaran
            
            # Get jumlah from the row if it exists, otherwise use running_balance
            jumlah = float(row["Jumlah"]) if "Jumlah" in row and pd.notna(row["Jumlah"]) else running_balance
            
            # Create transaction
            transaction = {
                "tanggal": row["Tanggal"].strftime("%Y-%m-%d")
                if pd.notna(row["Tanggal"])
                else "",
                "uraian": str(row["Uraian"]) if pd.notna(row["Uraian"]) else "",
                "penerimaan": penerimaan,
                "pengeluaran": pengeluaran,
                "jumlah": jumlah,
                "saldo_berjalan": running_balance  # Add running balance
            }
            transactions.append(transaction)

        except Exception as e:
            print(f"Error processing row {_ + 1}: {str(e)}")
            continue

    return transactions


def _create_excel_file(data: TemplateData, output_path: str) -> None:
    """
    Create an Excel file from the template data with dynamic columns.
    
    The Excel file will have consistent formatting with the template, including:
    - Proper number formatting for currency values
    - Auto-adjusted column widths
    - Styled header row with filters
    - Frozen header row
    - Consistent column ordering
    """
    try:
        print("\n=== Creating Excel File ===")
        print("Output path: " + output_path)
        print("Number of accounts: " + str(len(data.accounts) if data.accounts else 0))

        if not data.accounts:
            print("Warning: No accounts found in data")
            # Create a new workbook and save it
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Message"])
            ws.append(["No data available"])
            wb.save(output_path)
            return

        # Create a new workbook
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Remove the default sheet created by openpyxl
        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])

        for account in data.accounts:
            try:
                print("\nProcessing account: " + account.name)
                print(
                    "Number of transactions: "
                    + str(len(account.transactions) if account.transactions else 0)
                )

                # Create a DataFrame with all transactions
                rows = []
                running_balance = 0.0
                
                for tx in account.transactions or []:
                    # Calculate total penerimaan and pengeluaran for this transaction
                    total_penerimaan = sum((tx.penerimaan or {}).values()) if hasattr(tx, 'penerimaan') else 0.0
                    total_pengeluaran = sum((tx.pengeluaran or {}).values()) if hasattr(tx, 'pengeluaran') else 0.0
                    
                    # Calculate running balance
                    running_balance = running_balance + total_penerimaan - total_pengeluaran
                    
                    row = {
                        "Tanggal": getattr(tx, "tanggal", ""),
                        "Uraian": getattr(tx, "uraian", ""),
                        "Jumlah": getattr(tx, "jumlah", 0.0),
                        "Saldo Berjalan": running_balance,  # Use the calculated running balance
                    }

                    # Add penerimaan columns
                    if hasattr(tx, "penerimaan") and tx.penerimaan:
                        for key, value in (tx.penerimaan or {}).items():
                            row[f"Penerimaan_{key}"] = value

                    # Add pengeluaran columns
                    if hasattr(tx, "pengeluaran") and tx.pengeluaran:
                        for key, value in (tx.pengeluaran or {}).items():
                            row[f"Pengeluaran_{key}"] = value

                    rows.append(row)

                if not rows:
                    print(f"Warning: No transactions for account {account.name}")
                    ws = wb.create_sheet(title=account.name[:31])
                    ws.append(["Message"])
                    ws.append([f"No transactions for {account.name}"])
                    continue

                # Create DataFrame
                df = pd.DataFrame(rows)

                # Reorder columns: Tanggal, Uraian, Penerimaan_*, Pengeluaran_*, Saldo, Saldo Berjalan
                columns_order = ["Tanggal", "Uraian"]
                columns_order.extend(
                    sorted([col for col in df.columns if col.startswith("Penerimaan_")])
                )
                columns_order.extend(
                    sorted([col for col in df.columns if col.startswith("Pengeluaran_")])
                )
                if "Jumlah" in df.columns:
                    columns_order.append("Jumlah")
                if "Saldo Berjalan" in df.columns:
                    columns_order.append("Saldo Berjalan")

                # Keep only columns that exist in the DataFrame
                columns_order = [col for col in columns_order if col in df.columns]
                df = df[columns_order]

                # Create a new worksheet for this account
                sheet_name = account.name[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                
                # Define fills for different header types
                default_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                penerimaan_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")  # Light green
                pengeluaran_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
                
                # Define border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Write the header row
                ws.append(columns_order)
                
                # Write the data rows first
                for _, row in df.iterrows():
                    ws.append(row.tolist())
                
                # Now apply header formatting after all data is written
                for col_num, column_title in enumerate(columns_order, 1):
                    cell = ws.cell(row=1, column=col_num)
                    if str(column_title).startswith('Penerimaan_'):
                        cell.fill = penerimaan_fill
                    elif str(column_title).startswith('Pengeluaran_'):
                        cell.fill = pengeluaran_fill
                    else:
                        cell.fill = default_fill
                    
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                
                # Format data cells
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns_order)):
                    for cell in row:
                        cell.border = thin_border
                        # Format numeric cells (all except first two columns)
                        if cell.column_letter not in ['A', 'B'] and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0_);(#,##0)'
                
                # Set column widths
                for idx, column in enumerate(columns_order, 1):
                    col_letter = get_column_letter(idx)
                    if column == 'Uraian':
                        ws.column_dimensions[col_letter].width = 40
                    elif column == 'Tanggal':
                        ws.column_dimensions[col_letter].width = 12
                    elif column == 'Saldo Berjalan':
                        ws.column_dimensions[col_letter].width = 15
                    else:
                        ws.column_dimensions[col_letter].width = 18
                
                # Freeze the header row
                ws.freeze_panes = 'A2'
                
                # Add filters to the header row
                ws.auto_filter.ref = f"A1:{get_column_letter(len(columns_order))}1"

            except Exception as e:
                print(
                    f"Error processing account {getattr(account, 'name', 'unknown')}: {str(e)}"
                )
                import traceback
                traceback.print_exc()
                continue
        
        # If no worksheets were created, add a default one
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet("Data")
            ws.append(["Message"])
            ws.append(["No valid data to export"])
        
        # Save the workbook
        wb.save(output_path)
        print(f"Successfully saved Excel file to {output_path}")

    except Exception as e:
        print(f"\n!!! Error in _create_excel_file: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
